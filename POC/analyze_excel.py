"""
Analyze the Quannex Excel file structure and formulas
"""
import openpyxl
import json
import sys

def analyze_excel(file_path):
    """Analyze Excel file and extract structure, formulas, and data"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)

        analysis = {
            'sheets': [],
            'total_sheets': len(wb.sheetnames)
        }

        print(f"\n{'='*80}")
        print(f"EXCEL FILE ANALYSIS: {file_path}")
        print(f"{'='*80}\n")
        print(f"Total Sheets: {len(wb.sheetnames)}\n")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column

            sheet_info = {
                'name': sheet_name,
                'max_row': max_row,
                'max_col': max_col,
                'formulas': [],
                'named_ranges': [],
                'key_cells': []
            }

            print(f"\n{'─'*80}")
            print(f"SHEET: {sheet_name}")
            print(f"{'─'*80}")
            print(f"Dimensions: {max_row} rows × {max_col} columns\n")

            # Extract formulas and key cells
            formulas_found = []
            named_cells = []

            for row in ws.iter_rows(min_row=1, max_row=min(max_row, 100), min_col=1, max_col=min(max_col, 50)):
                for cell in row:
                    if cell.value:
                        # Check if it's a formula
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas_found.append({
                                'cell': cell.coordinate,
                                'formula': cell.value
                            })

                        # Check for key identifiers (headers, labels)
                        if isinstance(cell.value, str):
                            value_lower = cell.value.lower()
                            if any(keyword in value_lower for keyword in ['face', 'kpi', 'energy', 'coherence', 'pentagram', 'breath', 'ratio', 'axis', 'phi', 'golden', 'element', 'ball', 'pillar']):
                                named_cells.append({
                                    'cell': cell.coordinate,
                                    'value': cell.value
                                })

            # Print first 20 rows of data to understand structure
            print("First 20 rows preview:")
            print("-" * 120)
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                # Convert row to list and truncate very long values
                row_values = []
                for val in row[:15]:  # First 15 columns
                    if val is None:
                        row_values.append('')
                    elif isinstance(val, str):
                        row_values.append(val[:40] if len(val) > 40 else val)
                    else:
                        row_values.append(str(val)[:40])

                print(f"Row {row_idx:3d}: {' | '.join(str(v) for v in row_values)}")

            if formulas_found:
                print(f"\nFormulas found: {len(formulas_found)}")
                print("Sample formulas (first 10):")
                for formula in formulas_found[:10]:
                    print(f"  {formula['cell']}: {formula['formula'][:100]}")

            if named_cells:
                print(f"\nKey cells found: {len(named_cells)}")
                print("Sample key cells (first 20):")
                for cell in named_cells[:20]:
                    print(f"  {cell['cell']}: {cell['value']}")

            sheet_info['formulas'] = formulas_found
            sheet_info['key_cells'] = named_cells
            analysis['sheets'].append(sheet_info)

        print(f"\n{'='*80}")
        print("ANALYSIS COMPLETE")
        print(f"{'='*80}\n")

        return analysis

    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
        return None
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    file_path = r"C:\Users\murau\OneDrive\Stalinis kompiuteris\Spiral-Dashboard-First-Draft-(Quannex) (Recovered) (1).xlsx"

    # Install openpyxl if not available
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    analysis = analyze_excel(file_path)

    if analysis:
        # Save summary to JSON
        with open('excel_analysis_summary.json', 'w', encoding='utf-8') as f:
            json.dump(analysis, f, indent=2)
        print("\nAnalysis saved to excel_analysis_summary.json")
